from openpyxl import load_workbook,Workbook
from datetime import datetime
import re

def getData2019(filename):
    wb = load_workbook(filename=filename, data_only=True) # 或者数据，不是公式 用data_only
    sheet = wb.active
    count = sheet.max_row

    startRow=2  # 2nd row is the first data row
    dateRow=1  # column 1 is the date
    catRow=2    # column 2 is the category
    scoreRow=3   # column 3 is the score row
    numberRow=4  # column 4 is the number 
    isTechPilot=True

    data=[]
    row=startRow

    while row<count:
        # 读取第一行
        date=sheet.cell(row,dateRow).value
        cat = sheet.cell(row, catRow).value
        score = sheet.cell(row, scoreRow).value
        cat_sw=score
        row+=1
        if ('Tech Pilot draw' or 'tech-only draw') in date: # 如果有pilot，读后面4行
            isTechPilot=True
            scores=[]
            for i in range(3):
               score = sheet.cell(row, scoreRow).value 
               scores.append(score)
               row+=1
            cat_ig=scores[0]   
            cat_eesw=scores[1]   
            cat_eeig=scores[2]
            number=sheet.cell(row, numberRow).value
            row+=1
            
            cat_entry=''
            date=re.sub(r'\n',' ',date)
            date=date.replace('(Tech Pilot draw)','').strip()
            date=date.replace('(tech-only draw)','').strip()
            date=datetime.strptime(date,"%B %d, %Y")
            date=datetime.strftime(date,"%Y-%m-%d")
            data.append({"date":date,"sw":cat_sw,"ig":cat_ig,"elss":cat_entry,"eesw":cat_eesw,"eeig":cat_eeig,"number":number,'Tech Pilot':isTechPilot})
        else: #否则5行
            isTechPilot=False
            scores=[]
            for i in range(4):
               score = sheet.cell(row, scoreRow).value 
               scores.append(score)
               row+=1
            cat_ig=scores[0]   
            cat_entry=scores[1]   
            cat_eesw=scores[2]
            cat_eeig=scores[3] 
            number=sheet.cell(row, numberRow).value
            row+=1
            print(str(row)+':'+date)
            date=re.sub(r'\n',' ',date)
            date=date.replace('(Tech Pilot draw)','').strip()
            date=date.replace('(tech-only draw)','').strip()
            date=datetime.strptime(date,"%B %d, %Y")
            date=datetime.strftime(date,"%Y-%m-%d")
            data.append({"date":date,"sw":cat_sw,"ig":cat_ig,"elss":cat_entry,"eesw":cat_eesw,"eeig":cat_eeig,"number":number,'Tech Pilot':isTechPilot})
        

    return data

def getData2020(filename):
    wb = load_workbook(filename=filename, data_only=True) # 或者数据，不是公式 用data_only
    sheet = wb.active
    count = sheet.max_row

    startRow=2  # 2nd row is the first data row
    dateRow=1  # column 1 is the date
    catRow=3    # column 2 is the category
    scoreRow=4   # column 3 is the score row
    numberRow=2  # column 4 is the number 
    commentRow=5
    isTechPilot=True

    data=[]
    row=startRow

    while row<count:
        # 读取第一行
        date=sheet.cell(row,dateRow).value
        number=sheet.cell(row, numberRow).value
        cat = sheet.cell(row, catRow).value
        score = sheet.cell(row, scoreRow).value
        comment=sheet.cell(row, commentRow).value
        cat_sw=score
        row+=1
        if 'Tech Pilot draw' in comment: # 如果有pilot，读后面3行
            isTechPilot=True
            scores=[]
            for i in range(3):
               score = sheet.cell(row, scoreRow).value 
               scores.append(score)
               row+=1
            cat_ig=scores[0]   
            cat_eesw=scores[1]   
            cat_eeig=scores[2]
            cat_entry=''
            date=re.sub(r'\n',' ',date)
            date=datetime.strptime(date,"%B %d, %Y")
            date=datetime.strftime(date,"%Y-%m-%d")
            data.append({"date":date,"sw":cat_sw,"ig":cat_ig,"elss":cat_entry,"eesw":cat_eesw,"eeig":cat_eeig,"number":number,'Tech Pilot':isTechPilot})
        else: #否则4行
            isTechPilot=False
            scores=[]
            for i in range(4):
               score = sheet.cell(row, scoreRow).value 
               scores.append(score)
               row+=1
            cat_ig=scores[0]   
            cat_entry=scores[1]   
            cat_eesw=scores[2]
            cat_eeig=scores[3]   
            date=re.sub(r'\n',' ',date)
            date=datetime.strptime(date,"%B %d, %Y")
            date=datetime.strftime(date,"%Y-%m-%d")
            data.append({"date":date,"sw":cat_sw,"ig":cat_ig,"elss":cat_entry,"eesw":cat_eesw,"eeig":cat_eeig,"number":number,'Tech Pilot':isTechPilot})
        

    return data

def printData(dicts):
    for dict in dicts:
        for key in dict:
            print(key+'->'+str(dict[key]))

def save_dicts(dicts):
    book = Workbook()
    sheet = book.active

    sheet.append(['date','Skill Worker','International Graduate','Entry Level','EE-Skill Worker','EE-International Graduate','Invitation number','Tech Pilot?'])
    for dict in dicts:
        row=[]
        for key in dict:
            row.append(dict[key])
        print(row)    
        sheet.append(row)
    book.save('itas.xlsx')

#itas=getData2019('2017 ita.xlsx')
#itas=getData2020('200605 ita.xlsx')
#printData(itas)
#save_dicts(itas)

# 取得完数据后需要做的：
# 1. 扫描到的数据，跟excel里面最新的比较，如果是更新，把最新数据添加到现有excel中去
# 2. 统计当年ITA总数，打印： 总数，已经用完名额（regular的和Tech Pilot），剩余名额
# 3. 图形化表示数据走向
# 4. 计算出各项目最高分，最低分，和当年平均分
# 5. 计算出各项目历史年份的最高分，最低分，和该年的平均分




