# 事先做好数据整理，1. 写出所有员工的NOC代码 2. Rate栏将所有员工的工资转换为hourly rate.3. 日期都统一变成2020-12-31的格式
# 读取现有员工工资数据
from openpyxl import load_workbook,Workbook

def getData(filename):
    wb = load_workbook(filename=filename, data_only=True) # 或者数据，不是公式 用data_only
    sheet = wb.active
    count = sheet.max_row

    startRow=3  # 3rd row is the first data row
    hoursRow=10  # column 5 is the hours 
    nocRow=8    # column 8 is the noc code
    rateRow=9   # column 9 is the current rate 
    wages=[]
    for row in range(startRow, count+1):
        hours=sheet.cell(row,hoursRow).value
        noc = sheet.cell(row, nocRow).value  
        rate = sheet.cell(row, rateRow).value
        wages.append({"row":row,"noc":noc,"rate":rate,"hours":hours})
    return wages

# print(getData('Employee List.xlsx'))

# 根据NOC 找出lowest, median, highest
# 计算出Annual, 工资对应median的百分比，写入数据到新文件

def setData(wages,area_index):
    book = Workbook()
    sheet = book.active # 注意 这个不是函数 不能用（）
    rows=[]
    rows.append(['NOC','Hourly Rate','Hours/Week','Annual Wage','Lowest','Median','Highest','Percentage to Median'])
    for wage in wages:
        noc=wage.get("noc")
        rate=wage.get("rate")
        hours=wage.get("hours")
        annual=rate*hours*52
        
        wage_info=getNoc(noc,area_index)

        if wage_info[1]!=0:
            percentage=rate/wage_info[1]*100
        else:
            percentage='NA'
        
        row=[noc,rate,hours,annual,wage_info[0],wage_info[1],wage_info[2],percentage]
        rows.append(row)

    # 把所有行的数据，逐一append到excel的每个row里面
    for row in rows:
        sheet.append(row)
    book.save('noc.xlsx')

def getNoc(noc,area_index):
    if noc=='1123':
        return [22, 33,44]  
    else:
        return [11,12,15]

setData(getData('Employee List-Guangson.xlsx'),77)